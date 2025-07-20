import openpyxl
import numpy as np
import pandas as pd

# 设置 numpy 打印选项
np.set_printoptions(threshold=np.inf)


def load_data(path):
    """加载数据集"""
    # 加载 Excel 文件
    data_wb = openpyxl.load_workbook(path)

    # 获取各个工作表
    sheets = {
        'drug_s1': data_wb['drug_s1'],
        'target_s1': data_wb['target_s1'],
        'drug_s2': data_wb['drug_s2_test'],
        'target_s2': data_wb['target_s2_test'],
        'A': data_wb['A_test'],
    }

    # 生成所有的组合名称索引（用"-"分隔）
    data_pairs = []
    interaction_sheet = sheets['A']
    for row in range(2, interaction_sheet.max_row + 1):
        compound = interaction_sheet.cell(row, 1).value  # 第一列是化合物名称
        for col in range(2, interaction_sheet.max_column + 1):  # 遍历所有酶名称（跳过第一列）
            data = interaction_sheet.cell(1, col).value
            data_pairs.append(f"{compound}-{data}")

    # 加载各工作表数据
    def load_sheet_values(sheet):
        """从 Excel 工作表中提取数据并转换为 numpy 数组"""
        values = []
        for row in range(2, sheet.max_row + 1):  # 从第二行开始读取
            row_values = [sheet.cell(row, col).value for col in range(2, sheet.max_column + 1)]  # 从第二列开始读取
            values.append(row_values)
        return np.array(values)

    # 加载各工作表数据
    data_drug_s1_values = load_sheet_values(sheets['drug_s1'])
    data_target_s1_values = load_sheet_values(sheets['target_s1'])
    data_drug_s2_test_values = load_sheet_values(sheets['drug_s2'])
    data_target_s2_test_values = load_sheet_values(sheets['target_s2'])
    data_interaction_values = load_sheet_values(sheets['A'])


    # 返回所有数据
    return (
        data_drug_s1_values, data_target_s1_values,
        data_drug_s2_test_values, data_target_s2_test_values,
        data_interaction_values, data_pairs

    )

